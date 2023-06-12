import datetime
import json
import time

import urllib3
from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from sqlalchemy import create_engine, Column, String, JSON, Integer
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker
from webdriver_manager.chrome import ChromeDriverManager

urllib3.disable_warnings()
Base = declarative_base()


class DataEntry(Base):
    __tablename__ = 'data_entries'
    id = Column(Integer, primary_key=True)
    date = Column(String)
    data = Column(JSON)

def save_data_to_sqlite(data):
    date = data.pop("Date")
    data_json = json.dumps(data)
    engine = create_engine('sqlite:///data.db')
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    data_entry = DataEntry(date=date, data=data_json)
    session.add(data_entry)
    session.commit()
    session.close()



def scrape_data():
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    url = 'https://www.oree.com.ua/index.php/pricectr'
    driver.get(url)
    div_element = driver.find_element(By.XPATH,
                                      "//div[@class='dataTables_scrollBody']")
    next_day_data = div_element.find_elements(By.TAG_NAME, "tr")[-1]
    fields = next_day_data.find_elements(By.TAG_NAME, "td")
    data = {}

    for index, field in enumerate(fields):
        if index == 0:
            data["Date"] = field.text
        else:
            data[str(index) + ":00"] = field.text

    driver.quit()

    return data


def save_data_to_file(data_dict: dict):
    workbook = Workbook()
    sheet = workbook.active
    headers = list(data_dict.keys())

    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header

    values = list(data_dict.values())

    for col_num, value in enumerate(values, 1):
        sheet.cell(row=2, column=col_num).value = value

    filename = "data.xlsx"
    workbook.save(filename)


if __name__ == "__main__":
    current_time = datetime.datetime.now().strftime("%H:%M")
    for i in range(100):
        if current_time >= "13:00":
            data = scrape_data()
            while True:
                save_question = input(
                    "If you want to save data to database enter 'DB'\n"
                    "If you want to save data to xlsx-file enter 'F'\n")
                if save_question == "F":
                    save_data_to_file(data)
                    break
                elif save_question == "DB":
                    save_data_to_sqlite(data)
                    break

            print("Data was gathered successfully!")
            break
        time.sleep(300)
